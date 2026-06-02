"""Importador de cuadro comparativo comercial desde Excel."""

from __future__ import annotations

from io import BytesIO
from typing import Any
from uuid import UUID

from app.db.comparison_admin_repository import ComparisonAdminRepository


def _load_openpyxl():
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl requerido para importar Excel; instalar extra etl"
        ) from exc
    return openpyxl


async def import_commercial_xlsx(
    *,
    repo: ComparisonAdminRepository,
    subject_product_id: UUID,
    subject_product_name: str,
    file_bytes: bytes,
    source_document_id: UUID | None = None,
) -> dict[str, Any]:
    openpyxl = _load_openpyxl()
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if "comparativo comercial" in name.lower():
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    headers: list[str] = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val is not None and str(val).strip():
            headers.append(str(val).strip())
        elif headers:
            break

    rows: list[list[Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row_idx, col + 1).value for col in range(len(headers))]
        if any(v is not None and str(v).strip() for v in row):
            rows.append(row)

    set_id = await repo.get_or_create_set(subject_product_id)
    return await repo.import_commercial_sheet(
        set_id=set_id,
        subject_product_id=subject_product_id,
        subject_product_name=subject_product_name,
        headers=headers,
        rows=rows,
        source_document_id=source_document_id,
    )
