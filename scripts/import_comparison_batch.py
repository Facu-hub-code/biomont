#!/usr/bin/env python3
"""Importa y publica cuadros comparativos comerciales desde Excel.

Uso:
    DATABASE_URL=... python scripts/import_comparison_batch.py /path/to/file.xlsx ...
    DATABASE_URL=... python scripts/import_comparison_batch.py --dry-run /path/to/file.xlsx
"""

from __future__ import annotations

import argparse
import asyncio
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import UUID

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT / "services" / "common" / "src"))
sys.path.insert(0, str(_REPO_ROOT / "services" / "backoffice-api" / "src"))

import asyncpg  # noqa: E402
from pgvector.asyncpg import register_vector  # noqa: E402

from app.db.comparison_admin_repository import ComparisonAdminRepository  # noqa: E402
from biomont_common.db.pool import DatabasePool  # noqa: E402


@dataclass(frozen=True, slots=True)
class ImportJob:
    file_path: Path
    product_name: str
    subject_match_name: str
    sheet_hint: str | None = None
    header_row_hint: int | None = None


def _normalize_name(value: str) -> str:
    text = unicodedata.normalize("NFKD", value.strip().lower())
    return "".join(c for c in text.encode("ascii", "ignore").decode("ascii") if c.isalnum())


def _load_openpyxl():
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl requerido: pip install openpyxl") from exc
    return openpyxl


def _row_has_producto_header(values: list[Any]) -> bool:
    for val in values:
        if val is None:
            continue
        norm = _normalize_name(str(val))
        if norm in {"producto", "product"}:
            return True
    return False


def _pick_sheet(wb: Any, *, sheet_hint: str | None) -> str:
    if sheet_hint and sheet_hint in wb.sheetnames:
        return sheet_hint

    ranked: list[tuple[int, str]] = []
    for name in wb.sheetnames:
        lower = name.lower()
        score = 0
        if "comparativo comercial" in lower:
            score += 100
        elif "comparativo" in lower and "formula" not in lower and "fórmula" not in lower:
            score += 50
        elif "comparativo" in lower:
            score += 20

        ws = wb[name]
        header_row = None
        for row_idx in range(1, min(ws.max_row, 25) + 1):
            row_vals = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
            if _row_has_producto_header(row_vals):
                header_row = row_idx
                break
        if header_row is not None:
            score += 30 - header_row
        ranked.append((score, name))

    ranked.sort(reverse=True)
    best_score, best_name = ranked[0]
    if best_score <= 0:
        return wb.sheetnames[0]
    return best_name


def _find_header_row(ws: Any, *, header_row_hint: int | None) -> int:
    if header_row_hint is not None:
        return header_row_hint
    for row_idx in range(1, min(ws.max_row, 25) + 1):
        row_vals = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if _row_has_producto_header(row_vals):
            return row_idx
    return 1


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for header in headers:
        key = re.sub(r"[^a-z0-9]+", "_", header.strip().lower()).strip("_") or "column"
        count = seen.get(key, 0) + 1
        seen[key] = count
        out.append(header if count == 1 else f"{header} ({count})")
    return out


def parse_commercial_xlsx(
    file_bytes: bytes,
    *,
    sheet_hint: str | None = None,
    header_row_hint: int | None = None,
) -> tuple[list[str], list[list[Any]], str, int]:
    openpyxl = _load_openpyxl()
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    sheet_name = _pick_sheet(wb, sheet_hint=sheet_hint)
    ws = wb[sheet_name]
    header_row = _find_header_row(ws, header_row_hint=header_row_hint)

    header_start_col = 1
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is not None and str(val).strip():
            header_start_col = col
            break

    headers: list[str] = []
    for col in range(header_start_col, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is not None and str(val).strip():
            headers.append(str(val).strip())
        elif headers:
            break
    headers = _dedupe_headers(headers)

    rows: list[list[Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = [
            ws.cell(row_idx, header_start_col + col).value for col in range(len(headers))
        ]
        if any(v is not None and str(v).strip() for v in row):
            rows.append(row)

    wb.close()
    return headers, rows, sheet_name, header_row


async def _resolve_product(conn: asyncpg.Connection, product_name: str) -> tuple[UUID, str]:
    row = await conn.fetchrow(
        """
        SELECT id, name
        FROM public.products
        WHERE lower(name) = lower($1)
        LIMIT 1
        """,
        product_name,
    )
    if row is None:
        raise ValueError(f"producto_no_encontrado:{product_name}")
    return row["id"], row["name"]


async def _count_subject_rows(
    conn: asyncpg.Connection,
    *,
    set_id: UUID,
    published_version: int,
) -> int:
    return int(
        await conn.fetchval(
            """
            SELECT COUNT(*)
            FROM public.commercial_comparison_rows
            WHERE set_id = $1
              AND published_version = $2
              AND is_subject = true
            """,
            set_id,
            published_version,
        )
        or 0
    )


JOBS: list[ImportJob] = [
    ImportJob(
        Path("/Users/facundolorenzo/Downloads/COMPARATIVO COMERCIAL - AUMENTHA ATP NF.xlsx"),
        "Aumentha ATP NF",
        "AUMENTHA ATP NF",
        sheet_hint="1)",
        header_row_hint=3,
    ),
    ImportJob(
        Path("/Users/facundolorenzo/Downloads/COMPARATIVO COMERCIAL - GIGANTOL ADE.xlsx"),
        "Gigantol ADE",
        "GIGANTOL ADE",
    ),
    ImportJob(
        Path("/Users/facundolorenzo/Downloads/COMPARATIVO COMERCIAL - HEPATIN.xlsx"),
        "Hepatin",
        "Hepatin",
        sheet_hint="Hoja1",
        header_row_hint=3,
    ),
    ImportJob(
        Path("/Users/facundolorenzo/Downloads/COMPARATIVO COMERCIAL - TILOZONA.xlsx"),
        "Tilozona",
        "TILOZONA",
        sheet_hint="COMPARATIVO TILOSINA",
        header_row_hint=2,
    ),
    ImportJob(
        Path("/Users/facundolorenzo/Downloads/COMPARATIVO COMERCIAL PROTEGGO 3M y M.xlsx"),
        "Protego 3M",
        "Proteggo 3M",
        sheet_hint="Comparativo 3M y M",
    ),
    ImportJob(
        Path("/Users/facundolorenzo/Downloads/COMPARATIVO COMERCIAL TULABIOT.xlsx"),
        "Tulabiot",
        "TULABIOT",
    ),
    ImportJob(
        Path("/Users/facundolorenzo/Downloads/COMPARATIVO SEMENTAL (EX PM7,11 NF) 26.10.21.xlsx"),
        "Semental",
        "SEMENTAL",
        sheet_hint="COMPARATIVO",
    ),
    ImportJob(
        Path("/Users/facundolorenzo/Downloads/IMPERIA_Comparativo Comercial.xlsx"),
        "Imperia",
        "IMPERIA",
    ),
    ImportJob(
        Path("/Users/facundolorenzo/Downloads/KUAGULA_Comparativo Comercial.xlsx"),
        "Kuagula",
        "Kuagula",
        sheet_hint="Comparativo Perú ",
    ),
    ImportJob(
        Path("/Users/facundolorenzo/Downloads/MARVO 20_Comparativo Comercial.xlsx"),
        "Marvo 20",
        "MARVO 20",
    ),
    ImportJob(
        Path("/Users/facundolorenzo/Downloads/OPRURIX_Comparativo Comercial.xlsx"),
        "Opruix",
        "OPRURIX",
    ),
]


async def run_job(
    repo: ComparisonAdminRepository,
    conn: asyncpg.Connection,
    job: ImportJob,
    *,
    dry_run: bool,
) -> dict[str, Any]:
    if not job.file_path.exists():
        return {
            "product": job.product_name,
            "file": str(job.file_path),
            "status": "error",
            "message": "archivo_no_encontrado",
        }

    product_id, db_name = await _resolve_product(conn, job.product_name)
    raw = job.file_path.read_bytes()
    headers, rows, sheet_name, header_row = parse_commercial_xlsx(
        raw,
        sheet_hint=job.sheet_hint,
        header_row_hint=job.header_row_hint,
    )

    if not headers:
        return {
            "product": job.product_name,
            "file": job.file_path.name,
            "status": "error",
            "message": "sin_encabezados",
            "sheet": sheet_name,
        }
    if not rows:
        return {
            "product": job.product_name,
            "file": job.file_path.name,
            "status": "error",
            "message": "sin_filas",
            "sheet": sheet_name,
        }

    if dry_run:
        return {
            "product": job.product_name,
            "file": job.file_path.name,
            "status": "dry_run",
            "sheet": sheet_name,
            "header_row": header_row,
            "columns": len(headers),
            "rows": len(rows),
            "headers": headers[:6],
        }

    set_id = await repo.get_or_create_set(product_id)
    result = await repo.import_commercial_sheet(
        set_id=set_id,
        subject_product_id=product_id,
        subject_product_name=job.subject_match_name,
        headers=headers,
        rows=rows,
        source_document_id=None,
    )
    version = await repo.publish_set(set_id, published_by=None)
    subject_rows = await _count_subject_rows(conn, set_id=set_id, published_version=version)

    return {
        "product": db_name,
        "file": job.file_path.name,
        "status": "ok" if subject_rows > 0 else "warning",
        "sheet": sheet_name,
        "header_row": header_row,
        "imported_rows": result["imported_rows"],
        "columns": result["columns"],
        "gaps_created": result["gaps_created"],
        "published_version": version,
        "subject_rows": subject_rows,
        "message": None if subject_rows > 0 else "sin_fila_subject",
    }


async def main_async(files: list[Path] | None, *, dry_run: bool) -> int:
    database_url = os.environ.get("DATABASE_URL")
    if not database_url and not dry_run:
        raise SystemExit("DATABASE_URL requerida")

    jobs = JOBS
    if files:
        by_name = {job.file_path.name: job for job in JOBS}
        jobs = []
        for path in files:
            job = by_name.get(path.name)
            if job is None:
                raise SystemExit(f"Archivo no mapeado en JOBS: {path}")
            jobs.append(ImportJob(path, job.product_name, job.subject_match_name, job.sheet_hint, job.header_row_hint))

    pool = DatabasePool()
    if not dry_run:
        await pool.start()

    results: list[dict[str, Any]] = []
    try:
        if dry_run:
            conn = None
            repo = None
            for job in jobs:
                results.append(
                    {
                        "product": job.product_name,
                        "file": job.file_path.name,
                        "status": "dry_run_parse_only",
                    }
                )
                if job.file_path.exists():
                    headers, rows, sheet, header_row = parse_commercial_xlsx(
                        job.file_path.read_bytes(),
                        sheet_hint=job.sheet_hint,
                        header_row_hint=job.header_row_hint,
                    )
                    results[-1].update(
                        {
                            "sheet": sheet,
                            "header_row": header_row,
                            "columns": len(headers),
                            "rows": len(rows),
                        }
                    )
        else:
            async with pool.acquire() as conn:
                repo = ComparisonAdminRepository(pool)
                for job in jobs:
                    results.append(await run_job(repo, conn, job, dry_run=False))
    finally:
        if not dry_run:
            await pool.stop()

    ok = [r for r in results if r["status"] == "ok"]
    warn = [r for r in results if r["status"] == "warning"]
    err = [r for r in results if r["status"] == "error"]

    print("\n=== RESUMEN IMPORTACIÓN COMPARATIVOS ===\n")
    for r in results:
        status = r["status"].upper()
        line = f"[{status}] {r.get('product', '?')} <- {r.get('file', '?')}"
        if r.get("sheet"):
            line += f" | hoja={r['sheet']}"
        if r.get("imported_rows") is not None:
            line += f" | filas={r['imported_rows']} cols={r.get('columns')}"
        if r.get("published_version") is not None:
            line += f" | v{r['published_version']}"
        if r.get("subject_rows") is not None:
            line += f" | subject={r['subject_rows']}"
        if r.get("gaps_created"):
            line += f" | gaps={r['gaps_created']}"
        if r.get("message"):
            line += f" | {r['message']}"
        print(line)

    print(f"\nTotal: {len(ok)} OK, {len(warn)} warning, {len(err)} error")
    return 1 if err else 0


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("files", nargs="*", type=Path, help="Archivos xlsx (opcional)")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    raise SystemExit(asyncio.run(main_async(args.files or None, dry_run=args.dry_run)))


if __name__ == "__main__":
    main()
